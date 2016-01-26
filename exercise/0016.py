#-*- coding: utf-8 -*-
__author__ = "coolfire"
import json
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
i = 1
file_data = json.load(open("numbers.txt"))
for data in file_data:
	ws.cell("A%s" % i).value = "%s" % data[0]
	ws.cell("B%s" % i).value = "%s" % data[1]
	ws.cell("C%s" % i).value=  "%s" % data[2]
	i += 1
	if i > len(file_data):
		break
wb.save("numbers.xls")
