#-*- coding: utf-8 -*-
__author__ = "coolfire"
import json
from collections import OrderedDict
from openpyxl import Workbook
wb = Workbook()
ws1 = wb.active
i = 1
#load方法返回类文件对象
city_data = json.load(open ("city.txt"),encoding="utf8",object_pairs_hook=OrderedDict) #对返回的对象进行排序


for data in city_data.items():
	ws1.cell('A%s' % (i)).value = "%s" % data[0]
	ws1.cell('B%s' % (i)).value = '%s' % data[1]
	i += 1
	if i > len(city_data):
		break
wb.save("city.xls")


