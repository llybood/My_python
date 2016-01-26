#-*- coding: utf-8 -*-
__author__ = "coolfire"

import ConfigParser
import urllib
import urllib2
import MySQLdb
import os
import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

#编写快小递后台扩展功能,直接查询数据库获取一些数据信息,并按照一定规则来处理
#所以我们在这里把它封装成一个类,然后逐渐添加方法,来扩展我们的功能

class kuaixiaodi_extend:
    def __init__(self):
        #读取配置文件
        cf = ConfigParser.ConfigParser()
        cf.read("api.conf")
        #初始化一些变量
        self.user_agent="Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36"
        self.header={ "User_Agent" : self.user_agent }
        #手机号码归属地查询api接口地址
        self.ownership_api=cf.get("api","ownership_api")
        #数据库变量
        self.host = cf.get("db","db_host")
        self.user = cf.get("db","db_user")
        self.port = int(cf.get("db","db_port"))
        self.password = cf.get("db","db_password")

    #获取数据库数据,并且储存在txt文件中
    def Get_mysqldata(self,database,sql):
        phone_file=open("phone.txt","a")
        try:
            #连接数据库,获取全部数据
            conn = MySQLdb.connect(host=self.host,user=self.user,port=self.port,passwd=self.password,db=database)
            cur = conn.cursor()
            cur.execute(sql)
            rows = cur.fetchall()
            cur.close()
            #保存到文件
            for row in rows:
                phone_file.write(row[0]+"\n")
            phone_file.close()

        except MySQLdb.Error,e:
            print "Mysql Error %d: %s" % (e.args[0],e.args[1])

    #查询手机号码归属地
    def Inquire_phoneownership(self,phone):
        url = self.ownership_api + phone
        request = urllib2.Request(url,headers=self.header)
        try:
            response = urllib2.urlopen(request,timeout=5)
            #读取页面内容后,需要先以GBK解码,然后再以utf-8编码
            page_data = response.read().decode("gbk").encode("utf-8")
            #print page_data
            pattern = re.compile(r".*carrier:'(.+?)'")
            m = pattern.search(page_data)
            if m:
                print m.group(1)
                return m.group(1)
            else:
                return "None"
        except urllib2.URLError,e:
            if hasattr(e,"reason"):
                print "连接失败,失败原因: %s" % e.reason
                return None

    #批量查询手机号码归属地
    #传入手机号码文件,文件第一列为手机号码
    def Batch_inquire_phone(self,file_path):
        wb = Workbook()
        ws1 = wb.active
        i=1
        for phone in open(file_path):
            ownership=self.Inquire_phoneownership(phone)
            ws1.cell('A%s' % (i)).value =  '%s' % phone
            ws1.cell('B%s' % (i)).value =  '%s' % ownership
            i+=1
        wb.save('phone.xlsx')


    def Get_update_sql(self,database):
        #批量修改数据库记录的手机号码
        phone_list = ['15210684468','13810411458','18001289360','13401019783','15011239564','13311332800','15010589936','18610581904','15210149221','15600123661','13910558080','13381030298','18010401012','15201059291']
        i = 0
        try:
            conn = MySQLdb.connect(host=self.host,user=self.user,port=self.port,passwd=self.password,db=database)
            cur = conn.cursor()
            sql = 'select distinct phone from sms_records where sendStatus=0 order by id limit 5000'
            cur.execute(sql)
            rows = cur.fetchall()
            #for phone in phone_list:
            #   if phone in rows:
            #        print phone + "True"
            for num in range(0,5000,350):
                sql1 = 'select id from sms_records where phone=' + str(rows[num][0])
                cur.execute(sql1)
                id_number = cur.fetchone()
                print 'update sms_records set phone=' + phone_list[i] +  ' where id=' + str(id_number[0])
                i += 1
                if i >= 14:
                    break
            cur.close()
        except MySQLdb.Error,e:
            print "Mysql Error %d: %s" % (e.args[0],e.args[1])










test=kuaixiaodi_extend()
#test.Get_mysqldata("kuaiyou","select phone from t_ad_active")
#test.Inquire_phoneownership("15010589936")
test.Batch_inquire_phone("phone.txt")
#test.Get_update_sql("kuaiyou")


